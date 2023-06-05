import os

from openpyxl.workbook import Workbook
from selenium import webdriver
from bs4 import BeautifulSoup
import re
import datetime
from openpyxl import load_workbook
import json
import string
import sys

excelName = "apks.xlsx"
driver = webdriver.Chrome()  # 初始化Chrome浏览器驱动

type = "poker"

# 已经出包
slots_whitelist = ["com.iyhaofhvkm.phfmnvtrek", "com.fhgni1nd7fnr1im.cjenfinhrd", "com.slotsmachine.fcujoyfun",
                   "com.dnim.brazil", "pac.brazil.vamos", "com.slotsfavorites.slots.android",
                   "com.seersuckerbawqpx.kentuckykmyfu", "com.superluckycasino.viponly.slots.vegas.android.free",
                   "com.and.kk.jackpot.storm.game", "com.byte3d.poker", "com.giletech.pokeralfie", "com.emagssob.big2",
                   "com.arctomysjlmyg.rearrivelzvhoqy", "com.superluckycasino.nolimit.slots.vegas.android.free",
                   "app.bigwincasino.android2022", "com.bigwinlab.vegasslots", "com.slots.best.android",
                   "com.fuoops.colorslotsspin", "com.duksel.luckyspin.slots.free",
                   "com.ninefeet.slots.real.money.win.cash",
                   "com.spin.game.zhuan", "com.duksel.spinwin.slots.free", "com.zeuwnvfe.coeuse",
                   "com.pop.million.slots.gp",
                   "com.lq.spinlucky", "com.carnavalpraiaslots.sortedefelicidade.casinook890", "com.ppeihad.cient",
                   "com.mastergold.slots.kt", "com.prestigegame.egyptiangold", "com.dragonslots.gold",
                   "com.duksel.luckandwin.slots.free", "com.crossfield.slotsvegas", "com.ooppyey.supers",
                   "game.iuozxcpqeqw.xcpozcipix", "den.gogame.menbs", "com.funny.masterefun.slots3127",
                   "com.skdkooe.pprofo",
                   "com.prestigegame.stealinsheepfree", "com.foxcubgames.wolfbonuscasino", "com.od4n6usa.ucpd0371",
                   "com.acemegame.luckyslot", "naga.lucky.com", "com.slots.best.android", "net.flysher.rockncash",
                   "com.sagafun.casinoslots", "com.winning.jackpot.slots.casino.mania.slotmachines.free",
                   "com.sagafun.bingoage", "com.zariba.skillslots", "casino.games.slot.machines.free",
                   "com.gamelion.cleopatra", "air.com.wildtripleslots", "com.fibromyitisyhaikme.antigraphyhleipdg",
                   "com.ddy.slotgame.bingo.ddbingo", "com.tfunslots.cashcarnival.android", "com.huuuge.billionaires",
                   "com.huuuge.winter", "com.realcash.casino.slots", "com.arcade.clashfish", "org.hhhhhh.fqzscocos",
                   "com.rise.fishes.slots", "com.funny_corp.game.animpicas", "com.softcompany.cleocatra.slot",
                   "com.softcompany.thedoghousemultihold.slot", "air.com.gamblershome.dogslots",
                   "com.mobileamusements.PetStorePuppiesSlotsFREE", "com.zentertain.vegasgrandslots",
                   "com.ldddgames.farmanimalslotmaniagame", "com.motiblooman.arabicslots", "com.casino.slots.crash",
                   "com.RocketGames.TripleDouble", "com.megaSelf.game.slotGirls_Go", "com.funcasino.game.mphotocas",
                   "com.animan_publishing.mini_casino_slots", "com.rolling.lucky.slots.coin", "com.mslots.game",
                   "com.crazino",
                   "com.gametoast.lucky.buffalo.free.slots", "com.vegas.casino.slotmachines",
                   "com.duksel.goldenageofegypt.slots.free", "com.ar.BlueTreasureSlots", "com.patrick.greenbackslots"]

# 已经爬取过的
slots_excludeList = ["com.flyingturtle.holymoly", "com.duksel.yinyangslots.free",
                     "com.gametoast.lucky.buffalo.free.slots",
                     "com.crashpot", "com.pandg.millionsslots", "com.crazino", "itembox.crayfish.action"
                     ]

# 已经出包
dice_whitelist = ["com.fatartdesign.rockdice", "com.cjdesign.mightydice", "com.dicedom.dicemerge.puzzlegame",
                  "game.burjastudios.jhandimunda", "com.sampartridge.DiceRoll", "com.ccp.rpgsimpledice",
                  "appinventor.ai_ansonsavage.DiceRole_Magic8Ball_Extension_Two_Die""com.com.tann.dice"]

# 已经爬取过的
dice_excludeList = []

poker_whitelist = ["com.verylucky.pocker", "com.dtopoker.app", "com.giletech.pokeralfie", "com.jackpotmobile.tcp",
                   "com.emagssob.capsasusun", "com.youdagames.pokerworld", "com.devgamesc.pokerchellv1",
                   "com.joyp.pokermastergp", "com.emagssob.big2", "com.slgames.poker", "eu.ganymede.omahahd",
                   "net.gamoz.wpt", "com.AND48.PokerPreFlopTrainer", "com.bigwinlab.texasholdempoker",
                   "com.megagames.game.vpokerbattle", "com.solitaire.lucky.classic.card.and", "com.dtopoker.app",
                   "com.thegioilaptrinh.poker", "com.redgame.three", "com.qublix.ultimatePoker",
                   "com.avector.itw.itwtexasholdemhd", "com.youdagames.pokerhands", "com.shoreo.pokerg",
                   "com.kirorogame.sotaplus", "com.tapinator.vpc.ddbp"]

poker_excludeList = []


def start(key):
    letters = string.ascii_uppercase
    # zm=['h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
    for letter in letters:
        pageUrl = f"https://apkcombo.com/zh/search/{key}-{letter.lower()}#gsc.tab=0&gsc.q={key}%20m&gsc.sort="
        print(f"开始 {pageUrl}")
        findItems(pageUrl)


def findItems(url):
    soup = pullHtml(url)
    itemList = soup.findAll(class_="l_item")  # 查找class为"l_item"的元素
    for element in itemList:
        description_span = element.find("span", class_="description")  # 在每个元素中查找class为"description"的span标签元素
        if description_span is not None:
            spans = description_span.find_all('span')
            first_text = spans[0].text.strip()
            third_text = spans[2].text.strip()
            downloadSize = first_text.replace(" ", "").replace("+", "")

            apkSize = third_text.replace(" B", "").replace("N/A", "1").replace("KB", "").replace("MB", "").replace(" ",
                                                                                                                   "")  # apk的大小

            downloadNum = convert_to_number(downloadSize)  # apk下载量
            if (downloadNum is None or downloadNum <= 1_000_000) and 0 < int(apkSize) < 80:
                href_content = element['href']
                detailUrl = f"https://apkcombo.com/{href_content}"  # 解析拼接 符合条件的 详情URL

                detailHtml = pullHtml(detailUrl)  # 爬取详情页
                infomation = detailHtml.find(class_="information-table")
                divs = infomation.find_all("div", class_="value")
                pkg = 'divs[4].find("a").text.strip()'

                if type == "slots":
                    if slots_whitelist.__contains__(pkg) or slots_excludeList.__contains__(pkg):
                        print(f"{pkg} 存在 下一个")
                        continue
                elif type == "dice":
                    if dice_whitelist.__contains__(pkg) or dice_excludeList.__contains__(pkg):
                        print(f"{pkg} 存在 下一个")
                        continue
                elif type == "poker":
                    if poker_whitelist.__contains__(pkg) or poker_excludeList.__contains__(pkg):
                        print(f"{pkg} 存在 下一个")
                        continue

                dateStr = divs[1].text.strip()

                try:
                    date = datetime.datetime.strptime(dateStr, "%Y年%m月%d日")  # 解析后得到更新日期
                except ValueError:
                    now = datetime.datetime.now()  # 获取当前时间
                    date = now - datetime.timedelta(days=365)  # 减去一年的时间间隔 当作合格

                if date.year == 2023:  # 筛选2023年以前的
                    # gpUrl = divs[4].find("a")['href']
                    # # gpHtml = pull("https://play.google.com/store/apps/details?id=cash.mania.free.vegas.casino.slotqs&ref=apkcombo.com")
                    # gpHtml = pull(gpUrl)
                    # print(gpHtml)
                    # error_section = gpHtml.find('div', {'id': 'error-section'})
                    # if error_section is not None:
                    #     error_text = error_section.text
                    #     print(error_text)
                    #     print(f"gp未上线 {gpUrl}")
                    # else:
                    #     print('没有找到id为error-section的元素')
                    #     print(f"gp OK {gpUrl}")

                    downloadBt = detailHtml.find("a", class_="button is-success is-fullwidth")
                    downloadHref = f"https://apkcombo.com/{downloadBt['href']}"  # 解析拼接后得到下载页面URL
                    downloadHtml = pullHtml(downloadHref)
                    # 显式等待，等待元素加载完成
                    # codeTag = WebDriverWait(driver, 10).until(
                    #     EC.presence_of_element_located((By.TAG_NAME, "code"))
                    # )
                    codeTag = downloadHtml.find("code")
                    if codeTag is not None:
                        cpuarm = codeTag.text.strip()
                        if cpuarm.__contains__("universal") or cpuarm.__contains__(
                                "arm64-v8a") or cpuarm.__contains__(
                            "unknown"):  # 筛选支持64位的
                            name = element.find("span", class_="name").text.strip()
                            gpUrl = divs[4].find("a")['href']
                            pkgName = divs[4].find("a").text.strip()
                            data = {
                                "name": f"{name}",
                                "apkSize": f"{apkSize}MB",
                                "installNum": f"{downloadNum}",
                                "pkgName": pkgName,
                                "updateDate": dateStr,
                                "gpAddress": gpUrl,
                                "apkDownloadPage": downloadHref,
                                "cpuSupport": cpuarm,
                            }
                            json_string = json.dumps(data)
                            save2Excel(json_string)
                        else:
                            print(f"cup架构{cpuarm} 不合格 ")
                    else:
                        print(f"这里不应该......{downloadHref}")
                        printFile(detailUrl)


def save2Excel(data):
    print(f"保存:{data}")
    jsonStr = json.loads(data)
    creatExcel()
    # 加载Excel文件
    workbook = load_workbook(excelName)
    # 获取工作表名列表
    sheet_names = workbook.sheetnames
    # 判断工作表是否存在

    sheet_name = type
    if sheet_name in sheet_names:
        print(f'{sheet_name} 存在 无需新建')
        new_worksheet = workbook[sheet_name]
    else:
        print(f'{sheet_name} 不存在 需要创建工作表')
        new_worksheet = workbook.create_sheet(sheet_name)
        # 设置一些数据
        new_worksheet.cell(row=1, column=1).value = "名字"
        new_worksheet.cell(row=1, column=2).value = "包大小"
        new_worksheet.cell(row=1, column=3).value = "安装数"
        new_worksheet.cell(row=1, column=4).value = "包名"
        new_worksheet.cell(row=1, column=5).value = "更新日期"
        new_worksheet.cell(row=1, column=6).value = "gp地址"
        new_worksheet.cell(row=1, column=7).value = "apk下载地址"
        new_worksheet.cell(row=1, column=8).value = "cpu支持架构"

    # 追加数据到新的工作表
    next_row = new_worksheet.max_row + 1
    new_worksheet.cell(row=next_row, column=1).value = jsonStr["name"]
    new_worksheet.cell(row=next_row, column=2).value = jsonStr["apkSize"]
    new_worksheet.cell(row=next_row, column=3).value = jsonStr["installNum"]
    new_worksheet.cell(row=next_row, column=4).value = jsonStr["pkgName"]
    new_worksheet.cell(row=next_row, column=5).value = jsonStr["updateDate"]
    new_worksheet.cell(row=next_row, column=6).value = jsonStr["gpAddress"]
    new_worksheet.cell(row=next_row, column=7).value = jsonStr["apkDownloadPage"]
    new_worksheet.cell(row=next_row, column=8).value = jsonStr["cpuSupport"]

    # 保存修改后的Excel文件
    workbook.save(excelName)


def creatExcel():
    if not os.path.exists(excelName):
        print("apkcombo does not exist.")
        # 创建一个Workbook对象
        workbook = Workbook()

        # 保存文件
        workbook.save(excelName)


def pullHtml(url):
    driver.get(url)  # 打开网页
    content = driver.page_source
    soup = BeautifulSoup(content, "html.parser")  # 解析HTML
    return soup


def convert_to_number(text):
    match = re.search(r'(\d+)\s*([KM])', text)
    if match:
        value = int(match.group(1))
        unit = match.group(2)
        if unit == 'K':
            value *= 1000
        elif unit == 'M':
            value *= 1000000
        return value
    else:
        return None


def printFile(txt):
    # 打开文件
    with open(f'{type}.txt', 'a') as f:
        # 重定向 print 函数到文件流
        sys.stdout = f

        # 输出到文件
        print(f'{type}--{txt}\n')

        # 恢复 print 函数到标准输出流
        sys.stdout = sys.__stdout__


if __name__ == '__main__':
    start(type)
