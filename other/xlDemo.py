import openpyxl as openpyxl
import xlrd
from openpyxl import Workbook


def read(path):
    excel = openpyxl.load_workbook(path)
    for i in excel.active.rows:
        print(i)


def write(path):
    excel = Workbook()
    table = excel.active
    table['A1'] = "as1111"
    table['A4'] = "a42222"
    table.cell(5, 5, "cell 5 ,5 ")
    # table.write(0, 2, "hello 0_2")
    excel.save(path)


def full(path):
    excel = openpyxl.load_workbook(path)
    print(excel.sheetnames)
    sheet = excel.active

    for column_index, cells in enumerate(sheet.columns):
        for row_index, cell in enumerate(cells):
            print(f"{cell}  column:{column_index + 1} row:{row_index + 1}")
            if str(cell.value) == 'None':
                cell.value = f"cell {column_index}_{row_index} "
            else:
                print(cell.value)
    excel.save(path)


def modify(path):
    # 获取 工作簿对象
    excel = Workbook()
    table = excel.active
    table.row_dimensions[5].height = 80
    # table.cell(5, 5).width = 80

    # print(table.cell(5, 5).value )
    excel.save(path)


def crateSheet(path):
    workBook = Workbook()
    workBook.create_sheet("newSheet")

    workBook.save(path)


if __name__ == '__main__':
    write("./excel write test.xlsx")
    # full("./excel write test.xlsx")
    # read("./excel write test.xlsx")
    modify("./excel write test.xlsx")
    # crateSheet("./excel write test.xlsx")
